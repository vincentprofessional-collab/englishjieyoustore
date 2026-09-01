#!/usr/bin/env python3
"""Build a source-preserving intermediate representation for senior-high files.

This stage is deliberately not a question parser.  It records document
structure, styles, tables, media relationships, PDF coordinates, and source
hashes so later question parsing can be audited without revisiting the source
file.  The source directory is read-only; ZIP members are optionally expanded
under a separate cache with traversal checks.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import mimetypes
import posixpath
import re
import shutil
import subprocess
import tempfile
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET


ROOT = Path("/Users/shidianjin/Documents/高考英语")
REPO = Path(__file__).resolve().parents[1]
AUDIT = REPO / "data" / "senior-high" / "audit"
IR_DIR = REPO / "data" / "senior-high" / "document-ir"
ZIP_CACHE = REPO / "data" / "senior-high" / "zip-cache"
INVENTORY = AUDIT / "inventory.csv"
CHECKPOINT = AUDIT / "document-ir.checkpoint.json"
TEXT_EXTENSIONS = {".doc", ".docx", ".pdf", ".pptx", ".xlsx", ".xls", ".zip"}
MEDIA_EXTENSIONS = {".mp3", ".mp4", ".jpg", ".jpeg", ".png", ".wav", ".m4a", ".aac", ".flac", ".ogg"}

W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PR = "http://schemas.openxmlformats.org/package/2006/relationships"
A = "http://schemas.openxmlformats.org/drawingml/2006/main"
NS = {"w": W, "r": R, "pr": PR, "a": A}


def digest_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def source_ref(row: dict[str, str]) -> dict[str, Any]:
    return {
        "sourceDocumentId": row["sha256"],
        "relativePath": row["source_relpath"],
        "sha256": row["sha256"],
        "sizeBytes": int(row["size_bytes"]),
        "extension": row["extension"],
    }


def text_from_runs(parent: ET.Element) -> tuple[str, list[dict[str, Any]]]:
    runs: list[dict[str, Any]] = []
    for run in parent.findall("w:r", NS):
        parts: list[str] = []
        for node in run.iter():
            if node.tag == f"{{{W}}}t":
                parts.append(node.text or "")
            elif node.tag == f"{{{W}}}tab":
                parts.append("\t")
            elif node.tag in {f"{{{W}}}br", f"{{{W}}}cr"}:
                parts.append("\n")
        text = "".join(parts)
        if not text:
            continue
        rpr = run.find("w:rPr", NS)
        marks: list[str] = []
        if rpr is not None:
            for name in ("b", "i", "u", "vertAlign", "strike"):
                if rpr.find(f"w:{name}", NS) is not None:
                    marks.append(name)
        runs.append({"text": text, "marks": marks})
    return "".join(run["text"] for run in runs), runs


def relationship_targets(archive: zipfile.ZipFile, rel_path: str) -> dict[str, str]:
    try:
        root = ET.fromstring(archive.read(rel_path))
    except (KeyError, ET.ParseError):
        return {}
    result: dict[str, str] = {}
    base_dir = posixpath.dirname(rel_path).replace("/_rels", "")
    for relationship in root.findall("pr:Relationship", NS):
        target = relationship.attrib.get("Target", "")
        if target.startswith("/"):
            target = target.lstrip("/")
        else:
            target = posixpath.normpath(posixpath.join(base_dir, target))
        result[relationship.attrib.get("Id", "")] = target
    return result


def paragraph_block(paragraph: ET.Element, image_targets: dict[str, str]) -> dict[str, Any]:
    text, runs = text_from_runs(paragraph)
    p_style = paragraph.find("w:pPr/w:pStyle", NS)
    style = p_style.attrib.get(f"{{{W}}}val", "") if p_style is not None else ""
    images: list[str] = []
    for blip in paragraph.findall(".//a:blip", NS):
        embed = blip.attrib.get(f"{{{R}}}embed")
        if embed and embed in image_targets:
            images.append(image_targets[embed])
    block: dict[str, Any] = {
        "type": "heading" if style.lower().startswith("heading") else "paragraph",
        "text": text,
        "runs": runs,
        "style": style,
    }
    if images:
        block["imageTargets"] = images
    return block


def cell_blocks(cell: ET.Element, image_targets: dict[str, str]) -> list[dict[str, Any]]:
    blocks: list[dict[str, Any]] = []
    for child in list(cell):
        if child.tag == f"{{{W}}}p":
            block = paragraph_block(child, image_targets)
            if block["text"] or block.get("imageTargets"):
                blocks.append(block)
        elif child.tag == f"{{{W}}}tbl":
            blocks.append(table_block(child, image_targets))
    return blocks


def table_block(table: ET.Element, image_targets: dict[str, str]) -> dict[str, Any]:
    rows: list[list[list[dict[str, Any]]]] = []
    for row in table.findall("w:tr", NS):
        cells: list[list[dict[str, Any]]] = []
        for cell in row.findall("w:tc", NS):
            cells.append(cell_blocks(cell, image_targets))
        rows.append(cells)
    return {"type": "table", "rows": rows}


def parse_docx(path: Path, document_id: str) -> dict[str, Any]:
    blocks: list[dict[str, Any]] = []
    assets: list[dict[str, Any]] = []
    with zipfile.ZipFile(path) as archive:
        image_targets = relationship_targets(archive, "word/_rels/document.xml.rels")
        for name in sorted(member for member in archive.namelist() if member.startswith("word/media/")):
            data = archive.read(name)
            assets.append({
                "assetId": f"{document_id}:{name}",
                "kind": "image",
                "packagePath": name,
                "sha256": hashlib.sha256(data).hexdigest(),
                "sizeBytes": len(data),
                "mimeType": mimetypes.guess_type(name)[0] or "application/octet-stream",
            })
        root = ET.fromstring(archive.read("word/document.xml"))
        body = root.find("w:body", NS)
        if body is None:
            raise ValueError("DOCX has no document body")
        for child in list(body):
            if child.tag == f"{{{W}}}p":
                block = paragraph_block(child, image_targets)
                if block["text"] or block.get("imageTargets"):
                    blocks.append(block)
            elif child.tag == f"{{{W}}}tbl":
                blocks.append(table_block(child, image_targets))
    return {"blocks": blocks, "assets": assets, "paragraphCount": sum(block["type"] in {"paragraph", "heading"} for block in blocks), "tableCount": sum(block["type"] == "table" for block in blocks)}


def parse_doc(path: Path, document_id: str) -> dict[str, Any]:
    with tempfile.TemporaryDirectory(prefix="senior-high-doc-") as temp_dir:
        output_dir = Path(temp_dir)
        command = ["/Users/shidianjin/.cache/codex-runtimes/codex-primary-runtime/dependencies/bin/override/soffice", "--headless", "--convert-to", "docx", "--outdir", str(output_dir), str(path)]
        result = subprocess.run(command, capture_output=True, timeout=120, check=False)
        converted = output_dir / f"{path.stem}.docx"
        if result.returncode != 0 or not converted.exists():
            raise RuntimeError(f"soffice conversion failed: exit_{result.returncode}")
        parsed = parse_docx(converted, document_id)
        parsed["conversion"] = "soffice-doc-to-docx"
        return parsed


def parse_pdf(path: Path) -> dict[str, Any]:
    import fitz  # type: ignore

    pages: list[dict[str, Any]] = []
    text_chars = 0
    with fitz.open(path) as document:
        for page_number, page in enumerate(document, start=1):
            blocks: list[dict[str, Any]] = []
            for item in page.get_text("blocks"):
                text = item[4] or ""
                if text.strip():
                    text_chars += len(text)
                    blocks.append({"bbox": [round(float(value), 2) for value in item[:4]], "text": text})
            pages.append({"page": page_number, "width": page.rect.width, "height": page.rect.height, "blocks": blocks, "imageCount": len(page.get_images(full=True))})
    return {"pages": pages, "pageCount": len(pages), "textCharacterCount": text_chars, "needsOcr": text_chars == 0}


def parse_pptx(path: Path) -> dict[str, Any]:
    slides: list[dict[str, Any]] = []
    with zipfile.ZipFile(path) as archive:
        names = sorted((name for name in archive.namelist() if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)), key=lambda name: int(re.search(r"(\d+)", name).group(1)))
        for number, name in enumerate(names, start=1):
            root = ET.fromstring(archive.read(name))
            texts = [node.text or "" for node in root.findall(".//a:t", NS)]
            slides.append({"slide": number, "text": " ".join(texts), "paragraphs": [value for value in texts if value.strip()]})
    return {"slides": slides, "slideCount": len(slides)}


def parse_xlsx(path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook  # type: ignore

    workbook = load_workbook(path, read_only=True, data_only=False)
    sheets: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        rows: list[list[dict[str, Any]]] = []
        for row in sheet.iter_rows():
            cells = []
            for cell in row:
                if cell.value is not None:
                    cells.append({"coordinate": cell.coordinate, "value": str(cell.value), "dataType": cell.data_type})
            if cells:
                rows.append(cells)
        sheets.append({"sheet": sheet.title, "rows": rows})
    workbook.close()
    return {"sheets": sheets, "sheetCount": len(sheets)}


def safe_zip_member(name: str) -> bool:
    normalized = posixpath.normpath(name.replace("\\", "/"))
    return bool(normalized and normalized != "." and not normalized.startswith("../") and normalized != ".." and not normalized.startswith("/"))


def expand_zip(path: Path, document_id: str) -> tuple[list[dict[str, Any]], str]:
    destination = ZIP_CACHE / document_id
    destination.mkdir(parents=True, exist_ok=True)
    members: list[dict[str, Any]] = []
    with zipfile.ZipFile(path) as archive:
        for info in archive.infolist():
            safe = safe_zip_member(info.filename)
            record: dict[str, Any] = {"name": info.filename, "sizeBytes": info.file_size, "safe": safe}
            if safe and not info.is_dir():
                target = (destination / Path(info.filename)).resolve()
                if destination.resolve() not in target.parents:
                    raise ValueError(f"ZIP traversal target: {info.filename}")
                target.parent.mkdir(parents=True, exist_ok=True)
                if not target.exists() or target.stat().st_size != info.file_size:
                    with archive.open(info) as source, target.open("wb") as output:
                        shutil.copyfileobj(source, output)
                record["virtualSource"] = str(target.relative_to(REPO))
                record["sha256"] = digest_file(target)
            members.append(record)
    return members, str(destination.relative_to(REPO))


def parse_zip(path: Path, document_id: str, expand: bool) -> dict[str, Any]:
    with zipfile.ZipFile(path) as archive:
        members = [{"name": info.filename, "sizeBytes": info.file_size, "safe": safe_zip_member(info.filename)} for info in archive.infolist()]
    result: dict[str, Any] = {"members": members, "memberCount": len(members)}
    if expand:
        expanded, cache = expand_zip(path, document_id)
        result["expandedMembers"] = expanded
        result["cacheDirectory"] = cache
    return result


def parse_media(path: Path, document_id: str) -> dict[str, Any]:
    return {"assets": [{"assetId": document_id, "kind": "audio" if path.suffix.lower() in {".mp3", ".wav", ".m4a", ".aac", ".flac", ".ogg"} else "video" if path.suffix.lower() == ".mp4" else "image", "relativePath": path.name, "sha256": document_id, "sizeBytes": path.stat().st_size, "mimeType": mimetypes.guess_type(path.name)[0] or "application/octet-stream"}]}


def parse_source(path: Path, document_id: str, expand_zips: bool) -> tuple[str, dict[str, Any], str]:
    extension = path.suffix.lower()
    if extension == ".docx":
        return "docx-xml", parse_docx(path, document_id), "ok"
    if extension == ".doc":
        return "docx-xml-via-soffice", parse_doc(path, document_id), "ok"
    if extension == ".pdf":
        content = parse_pdf(path)
        return "pdf-text-coordinates", content, "needs_ocr" if content.get("needsOcr") else "ok"
    if extension == ".pptx":
        return "pptx-xml", parse_pptx(path), "ok"
    if extension in {".xlsx", ".xls"}:
        return "xlsx-cell", parse_xlsx(path), "ok"
    if extension == ".zip":
        return "zip-manifest", parse_zip(path, document_id, expand_zips), "ok"
    if extension in MEDIA_EXTENSIONS:
        return "media-manifest", parse_media(path, document_id), "ok"
    return "unsupported", {}, "unsupported"


def load_inventory(path: Path, selected_paths: list[str] | None, root: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    by_rel = {row["source_relpath"]: row for row in rows}
    if not selected_paths:
        return rows
    selected: list[dict[str, str]] = []
    for value in selected_paths:
        candidate = Path(value).expanduser()
        relative = candidate.resolve().relative_to(root).as_posix() if candidate.is_absolute() else candidate.as_posix()
        if relative not in by_rel:
            raise SystemExit(f"selected path is absent from inventory: {relative}")
        selected.append(by_rel[relative])
    return selected


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=ROOT)
    parser.add_argument("--inventory", type=Path, default=INVENTORY)
    parser.add_argument("--out-dir", type=Path, default=IR_DIR)
    parser.add_argument("--checkpoint", type=Path, default=CHECKPOINT)
    parser.add_argument("--path", action="append", dest="paths", help="Process selected absolute or root-relative files only.")
    parser.add_argument("--expand-zips", action="store_true")
    args = parser.parse_args()
    root = args.root.expanduser().resolve()
    rows = load_inventory(args.inventory.resolve(), args.paths, root)
    by_hash: dict[str, dict[str, Any]] = {}
    source_refs: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        source_refs.setdefault(row["sha256"], []).append(source_ref(row))
        by_hash.setdefault(row["sha256"], row)
    out_dir = args.out_dir.resolve()
    checkpoint_path = args.checkpoint.resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    checkpoint: dict[str, Any] = {}
    if checkpoint_path.exists():
        try:
            checkpoint = json.loads(checkpoint_path.read_text(encoding="utf-8")).get("records", {})
        except (OSError, json.JSONDecodeError, TypeError):
            checkpoint = {}
    processed = 0
    reused = 0
    statuses: dict[str, int] = {}
    for document_id, row in sorted(by_hash.items()):
        output = out_dir / f"{document_id}.json"
        previous = checkpoint.get(document_id)
        if previous and output.exists():
            reused += 1
            record = previous
            record["sourceRefs"] = source_refs[document_id]
        else:
            path = Path(row["source_file"])
            try:
                method, content, status = parse_source(path, document_id, args.expand_zips)
                record = {
                    "schemaVersion": 1,
                    "documentId": document_id,
                    "sourceRefs": source_refs[document_id],
                    "extension": row["extension"],
                    "extractionMethod": method,
                    "status": status,
                    "extractedAt": datetime.now(timezone.utc).isoformat(),
                    "content": content,
                }
                output.write_text(json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")
                processed += 1
            except Exception as error:  # keep one bad source from hiding the rest
                record = {
                    "schemaVersion": 1,
                    "documentId": document_id,
                    "sourceRefs": source_refs[document_id],
                    "extension": row["extension"],
                    "extractionMethod": "failed",
                    "status": "review_required",
                    "reason": f"{type(error).__name__}: {error}",
                    "extractedAt": datetime.now(timezone.utc).isoformat(),
                    "content": {},
                }
                output.write_text(json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")
                processed += 1
        checkpoint[document_id] = record
        statuses[record["status"]] = statuses.get(record["status"], 0) + 1
        if len(checkpoint) % 50 == 0:
            checkpoint_path.parent.mkdir(parents=True, exist_ok=True)
            checkpoint_path.write_text(json.dumps({"version": 1, "complete": False, "root": str(root), "records": checkpoint}, ensure_ascii=False, indent=2), encoding="utf-8")
    checkpoint_path.parent.mkdir(parents=True, exist_ok=True)
    checkpoint_path.write_text(json.dumps({"version": 1, "complete": True, "root": str(root), "records": checkpoint}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"root": str(root), "uniqueSources": len(by_hash), "processed": processed, "reused": reused, "statuses": statuses, "outputDir": str(out_dir), "checkpoint": str(checkpoint_path)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
