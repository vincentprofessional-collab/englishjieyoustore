#!/usr/bin/env python3
"""Extract source text locally, one byte-unique source at a time.

This is an audit stage, not a publisher.  It preserves the source path/hash
mapping and writes no files into the source directory.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import subprocess
import sys
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


DEFAULT_ROOT = Path("/Users/shidianjin/Documents/高考英语")
DEFAULT_AUDIT = Path("/Users/shidianjin/ielts-platform/data/senior-high/audit")
DEFAULT_TEXT = Path("/Users/shidianjin/ielts-platform/data/senior-high/source-text")
TEXT_EXTENSIONS = {".doc", ".docx", ".pdf", ".pptx", ".xlsx", ".xls", ".zip"}


def run_capture(command: list[str], timeout: int = 90) -> tuple[str, str]:
    try:
        result = subprocess.run(command, capture_output=True, timeout=timeout, check=False)
    except (OSError, subprocess.TimeoutExpired) as error:
        return "", f"{type(error).__name__}"
    stdout = result.stdout.decode("utf-8", errors="replace")
    stderr = result.stderr.decode("utf-8", errors="replace")
    if result.returncode != 0:
        return stdout, f"exit_{result.returncode}:{stderr[-300:]}"
    return stdout, "ok"


def extract_pptx(path: Path) -> tuple[str, str]:
    paragraphs: list[str] = []
    try:
        with zipfile.ZipFile(path) as archive:
            slide_names = sorted(name for name in archive.namelist() if re.match(r"ppt/slides/slide\d+\.xml$", name))
            for slide_name in slide_names:
                xml = archive.read(slide_name).decode("utf-8", errors="replace")
                texts = re.findall(r"<a:t[^>]*>(.*?)</a:t>", xml, flags=re.DOTALL)
                slide_text = " ".join(re.sub(r"<[^>]+>", "", value) for value in texts).strip()
                if slide_text:
                    paragraphs.append(f"[{slide_name}]\n{slide_text}")
    except (OSError, zipfile.BadZipFile) as error:
        return "", f"{type(error).__name__}"
    return "\n\n".join(paragraphs), "ok"


def extract_xlsx(path: Path) -> tuple[str, str]:
    try:
        # The bundled runtime has openpyxl; importing it here keeps the script
        # usable even when the system Python does not.
        from openpyxl import load_workbook  # type: ignore

        workbook = load_workbook(path, read_only=True, data_only=True)
        chunks: list[str] = []
        for sheet in workbook.worksheets:
            rows: list[str] = [f"[Sheet: {sheet.title}]"]
            for row in sheet.iter_rows(values_only=True):
                values = ["" if value is None else str(value) for value in row]
                if any(values):
                    rows.append("\t".join(values))
            chunks.append("\n".join(rows))
        workbook.close()
        return "\n\n".join(chunks), "ok"
    except Exception as error:  # a source-specific parse failure belongs in the audit
        return "", f"{type(error).__name__}"


def extract_zip_listing(path: Path) -> tuple[str, str]:
    try:
        with zipfile.ZipFile(path) as archive:
            entries = []
            for info in archive.infolist():
                entries.append(f"{info.filename}\t{info.file_size}")
            return "[ARCHIVE CONTENTS]\n" + "\n".join(entries), "ok_archive_listing"
    except (OSError, zipfile.BadZipFile) as error:
        return "", f"{type(error).__name__}"


def extract(path: Path) -> tuple[str, str]:
    extension = path.suffix.lower()
    if extension in {".doc", ".docx"}:
        return run_capture(["/usr/bin/textutil", "-convert", "txt", "-stdout", str(path)])
    if extension == ".pdf":
        return run_capture(["/opt/homebrew/bin/pdftotext", "-layout", str(path), "-"])
    if extension == ".pptx":
        return extract_pptx(path)
    if extension in {".xlsx", ".xls"}:
        return extract_xlsx(path)
    if extension == ".zip":
        return extract_zip_listing(path)
    return "", "not_text_source"


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fields = ["sha256", "canonical_source_file", "source_relpath", "extension", "size_bytes", "output_file", "status", "character_count", "line_count", "extracted_at"]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=DEFAULT_ROOT)
    parser.add_argument("--audit", type=Path, default=DEFAULT_AUDIT)
    parser.add_argument("--text-out", type=Path, default=DEFAULT_TEXT)
    args = parser.parse_args()
    audit = args.audit.resolve()
    text_out = args.text_out.resolve()
    inventory_path = audit / "inventory.csv"
    if not inventory_path.exists():
        print(f"missing inventory: {inventory_path}", file=sys.stderr)
        return 2
    text_out.mkdir(parents=True, exist_ok=True)
    checkpoint_path = audit / "text-extraction.checkpoint.json"
    checkpoint: dict[str, dict[str, Any]] = {}
    if checkpoint_path.exists():
        try:
            checkpoint = json.loads(checkpoint_path.read_text(encoding="utf-8")).get("records", {})
        except (OSError, json.JSONDecodeError, TypeError):
            checkpoint = {}
    with inventory_path.open(encoding="utf-8-sig", newline="") as handle:
        inventory = list(csv.DictReader(handle))
    by_hash: dict[str, dict[str, Any]] = {}
    for row in inventory:
        digest = row.get("sha256", "")
        if row.get("read_status") != "ok" or not digest:
            continue
        if digest not in by_hash or len(row["source_relpath"]) < len(by_hash[digest]["source_relpath"]):
            by_hash[digest] = row

    rows: list[dict[str, Any]] = []
    pending = 0
    for digest, row in sorted(by_hash.items()):
        path = Path(row["source_file"])
        extension = row["extension"]
        output = text_out / f"{digest}.txt"
        if digest in checkpoint and output.exists():
            result = checkpoint[digest]
        elif extension not in TEXT_EXTENSIONS:
            result = {
                "sha256": digest,
                "canonical_source_file": str(path),
                "source_relpath": row["source_relpath"],
                "extension": extension,
                "size_bytes": row["size_bytes"],
                "output_file": str(output),
                "status": "not_text_source",
                "character_count": 0,
                "line_count": 0,
                "extracted_at": datetime.now(timezone.utc).isoformat(),
            }
        else:
            pending += 1
            content, status = extract(path)
            if status.startswith("ok") and content:
                output.write_text(content.replace("\x00", ""), encoding="utf-8")
            result = {
                "sha256": digest,
                "canonical_source_file": str(path),
                "source_relpath": row["source_relpath"],
                "extension": extension,
                "size_bytes": row["size_bytes"],
                "output_file": str(output),
                "status": status if content or status != "ok" else "empty_text",
                "character_count": len(content),
                "line_count": content.count("\n") + (1 if content else 0),
                "extracted_at": datetime.now(timezone.utc).isoformat(),
            }
        checkpoint[digest] = result
        rows.append(result)
        if len(rows) % 100 == 0:
            checkpoint_path.write_text(json.dumps({"version": 1, "records": checkpoint}, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"checkpoint unique_sources={len(rows)} pending_processed={pending}")
    checkpoint_path.write_text(json.dumps({"version": 1, "complete": True, "records": checkpoint}, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(audit / "text-manifest.csv", rows)
    print(json.dumps({
        "unique_hashed_sources": len(by_hash),
        "text_manifest_rows": len(rows),
        "processed_now": pending,
        "extractable_outputs": sum(row["status"].startswith("ok") for row in rows),
        "non_text_sources": sum(row["status"] == "not_text_source" for row in rows),
        "output_dir": str(text_out),
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
